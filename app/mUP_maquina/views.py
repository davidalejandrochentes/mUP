from django.shortcuts import render, redirect, get_object_or_404
from .models import Maquina, MantenimientoMaquina, TipoMantenimientoMaquina, HorasParaAlerta
from .forms import MaquinaForm, MantenimientoMaquinaCorrectivoForm, MantenimientoMaquinaPreventivoForm, HorasParaAlertaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------------------------

@login_required
def maquina(request):
    horas_alerta = HorasParaAlerta.objects.first().horas
    alert = Maquina.objects.all()
    maquinas = Maquina.objects.filter(nombre__icontains=request.GET.get('search', ''))
    total_maquinas = len(maquinas)
    alertas = []
    for maquina in alert:
        horas_restantes = maquina.horas_restantes_mantenimiento()
        if horas_restantes <= horas_alerta:
            
            alertas.append({
                'maquina': maquina,
                'horas_restantes': horas_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['horas_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'maquinas': maquinas,
        'total_maquinas': total_maquinas,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'mUP_maquina/maquina.html', context)




@login_required
def alertas(request):
    horas_alert = get_object_or_404(HorasParaAlerta, id=1)
    if request.method == 'POST':
        alert_form = HorasParaAlertaForm(request.POST, instance=horas_alert) 
        if alert_form.is_valid():
            horas = alert_form.cleaned_data.get('horas')
            if horas < 1:
                return redirect('maquina_alertas')
            else:
                alert_form.save()
                return redirect('maquina_alertas')    
    else:
        alert_form = HorasParaAlertaForm(instance=horas_alert)
    horas_alerta = horas_alert.horas
    
    alert = Maquina.objects.filter(nombre__icontains=request.GET.get('search', ''))
    alertas = []
    for maquina in alert:
        horas_restantes = maquina.horas_restantes_mantenimiento()
        if horas_restantes <= horas_alerta:
            
            alertas.append({
                'maquina': maquina,
                'horas_restantes': horas_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['horas_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
        'alert_form': alert_form,
    }
    return render(request, 'mUP_maquina/alertas.html', context)




@login_required
def tabla_mantenimientos(request):
    maquinas = Maquina.objects.all()
    tipos_mantenimiento = TipoMantenimientoMaquina.objects.all()
    for maquina in maquinas:
        maquina.mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
    context = {
        'maquinas': maquinas,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'mUP_maquina/tablas.html', context)




@login_required
def crear_maquina(request):
    if request.method == 'GET':
        form = MaquinaForm()
        context = {
            'form': form
        }
        return render(request, 'mUP_maquina/nueva.html', context)
    if request.method == 'POST':
        form = MaquinaForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'mUP_maquina/nueva.html', context)
            else:
                # Manejo del archivo de imagen
                if 'imagen' in request.FILES:
                    form.instance.imagen = request.FILES['imagen']
                form.save()
                return redirect('maquina')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'mUP_maquina/nueva.html', context)




@login_required    
def detalles(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
        form = MaquinaForm(instance=maquina)
        context = {
            'maquina': maquina,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'mUP_maquina/detalles.html', context)
    
    if request.method == 'POST':
        maquina = get_object_or_404(Maquina, id=id)
        form = MaquinaForm(instance=maquina)
        form = MaquinaForm(request.POST, request.FILES, instance=maquina)

        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'maquina': maquina,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            else:
                form.save()
                mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
                context = {
                    'maquina': maquina,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                return render(request, 'mUP_maquina/detalles.html', context) 
        
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url)




@login_required
def eliminar(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    maquina.delete()
    return redirect ('maquina') 

# fin de vistas generales----------------------------------------------------------------------------------




@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)




@login_required
def mantenimientos_maquina_preventivo(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2) 
        mantenimientos = maquina.mantenimientomaquina_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'maquina': maquina,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'mUP_maquina/manteniminetos_preventivo.html', context)   




@login_required
def mod_mantenimineto_maquina_preventivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
        maquina = mantenimiento.maquina
        form_mant = MantenimientoMaquinaPreventivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'maquina': maquina,
        }
        return render(request, 'mUP_maquina/mod_mantenimineto_preventivo.html', context)
    
    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2) 
        maquina = mantenimiento.maquina
        form_mant = MantenimientoMaquinaPreventivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.maquina = maquina
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'imagen' in request.FILES:
                mantenimiento.imagen = request.FILES['imagen'] 
            mantenimiento.save()
            return redirect('mantenimientos_maquina_preventivo', id=maquina.id)
        else:
            context = {
                'form_mant': form_mant,
                'maquina': maquina,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'mUP_maquina/mod_mantenimineto_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_maquina_preventivo(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2) 
        form_mant = MantenimientoMaquinaPreventivoForm()
        context = {
            'form_mant': form_mant,
            'maquina': maquina,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html', context)
    
    if request.method == 'POST':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2) 
        form_mant = MantenimientoMaquinaPreventivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.maquina = maquina
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.partes_y_piezas = ""
            if 'imagen' in request.FILES:
                mantenimiento.imagen = request.FILES['imagen'] 
            mantenimiento.save()
            return redirect('mantenimientos_maquina_preventivo', id=maquina.id)
        else:
            context = {
                'form_mant': form_mant,
                'maquina': maquina,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def mantenimientos_maquina_correctivo(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1) 
        mantenimientos = maquina.mantenimientomaquina_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'maquina': maquina,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'mUP_maquina/manteniminetos_correctivo.html', context)   




@login_required
def mod_mantenimineto_maquina_correctivo(request, id):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
        maquina = mantenimiento.maquina
        form_mant = MantenimientoMaquinaCorrectivoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'maquina': maquina,
        }
        return render(request, 'mUP_maquina/mod_mantenimineto_correctivo.html', context)
    
    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1) 
        maquina = mantenimiento.maquina
        form_mant = MantenimientoMaquinaCorrectivoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.maquina = maquina
            mantenimiento.tipo = tipo_mantenimiento
            if 'imagen' in request.FILES:
                mantenimiento.imagen = request.FILES['imagen'] 
            mantenimiento.save()
            return redirect('mantenimientos_maquina_correctivo', id=maquina.id)
        else:
            context = {
                'form_mant': form_mant,
                'maquina': maquina,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'mUP_maquina/mod_mantenimineto_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




@login_required
def nuevo_mantenimineto_maquina_correctivo(request, id):
    if request.method == 'GET':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1) 
        form_mant = MantenimientoMaquinaCorrectivoForm()
        context = {
            'form_mant': form_mant,
            'maquina': maquina,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html', context)
    
    if request.method == 'POST':
        maquina = get_object_or_404(Maquina, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1) 
        form_mant = MantenimientoMaquinaCorrectivoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.maquina = maquina
            mantenimiento.tipo = tipo_mantenimiento
            
            if form_mant.cleaned_data['hr_maquina'] > maquina.horas_máquina_trabajada:
                form_mant.add_error('hr_maquina', 'Las horas de trabajo del mantenimiento no pueden ser mayores que las horas de trabajo de la máquina.')
                context = {
                    'form_mant': form_mant,
                    'maquina': maquina,
                    'tipo_mantenimiento': tipo_mantenimiento,
                }
                messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
                return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html', context)
            else:
                if 'imagen' in request.FILES:
                    mantenimiento.imagen = request.FILES['imagen'] 
                mantenimiento.save()
                return redirect('mantenimientos_maquina_correctivo', id=maquina.id)
        else:
            context = {
                'form_mant': form_mant,
                'maquina': maquina,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html', context)

    return HttpResponse("Method Not Allowed", status=405)




# descargas -----------------------------------------------------------------------------------

@login_required
def documento_general_mantenimientos_maquina(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoMaquina.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_maquinas_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_maquinas_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Maquina', 'Tipo', 'Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.maquina.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_preventivos_maquina(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 2

    maquina = get_object_or_404(Maquina, pk=id)
    mantenimientos = MantenimientoMaquina.objects.filter(maquina=maquina).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}.xlsx"'.format(maquina.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}.xlsx"'.format(maquina.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_correctivos_maquina(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 1

    maquina = get_object_or_404(Maquina, pk=id)
    mantenimientos = MantenimientoMaquina.objects.filter(maquina=maquina).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}_{}.xlsx"'.format(maquina.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}.xlsx"'.format(maquina.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response