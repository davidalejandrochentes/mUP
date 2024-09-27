from django.shortcuts import render, redirect, get_object_or_404
from .models import Vehiculo, MantenimientoVehiculo, TipoMantenimientoVehiculo, KmParaAlerta
from .forms import VehiculoForm, MantenimientoVehiculoForm, KmParaAlertaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.forms import AuthenticationForm

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------------------------

@login_required
def vehiculo(request):
    km_alerta = KmParaAlerta.objects.first().km
    alert = Vehiculo.objects.all()
    vehiculos = Vehiculo.objects.filter(marca__icontains=request.GET.get('search', ''))
    total_vehiculos = len(vehiculos)
    alertas = []
    for vehiculo in alert:

        km_restantes_mantenimiento_correctivo = vehiculo.km_restantes_mantenimiento_correctivo()
        if km_restantes_mantenimiento_correctivo <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_mantenimiento_correctivo,
            })

        # Cambio de filtro de aceite
        km_restantes_filtro_aceite = vehiculo.km_restantes_intervalo_cambio_filtro_aceite()
        if km_restantes_filtro_aceite <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aceite,
            })

        # Cambio de filtro de aire/combustible
        km_restantes_filtro_aire_combustible = vehiculo.km_restantes_intervalo_cambio_filtro_aire_combustible()
        if km_restantes_filtro_aire_combustible <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aire_combustible,
            })

        # Cambio de filtro de caja/corona
        km_restantes_filtro_caja_corona = vehiculo.km_restantes_intervalo_cambio_aceite_caja_corona()
        if km_restantes_filtro_caja_corona <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_caja_corona,
            })
    
    alertas_ordenadas = sorted(alertas, key=lambda x: x['km_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'vehiculos': vehiculos,
        'total_vehiculos': total_vehiculos,
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
    }
    return render(request, 'SGE_vehiculo/vehiculo.html', context)




@login_required
def alertas(request):
    km_alert = get_object_or_404(KmParaAlerta, id=1)
    if request.method == 'POST':
        alert_form = KmParaAlertaForm(request.POST, instance=km_alert)
        if alert_form.is_valid():
            km = alert_form.cleaned_data.get('km')
            if km < 1:
                return redirect('vehiculo_alertas')
            else:
                alert_form.save()
                return redirect('vehiculo_alertas')
    else:
        alert_form = KmParaAlertaForm(instance=km_alert)
    km_alerta = km_alert.km                   
    search_query = request.GET.get('search', '')
    vehiculos = Vehiculo.objects.filter(marca__icontains=search_query)
    alertas = []
    
    for vehiculo in vehiculos:

        km_restantes_mantenimiento_correctivo = vehiculo.km_restantes_mantenimiento_correctivo()
        if km_restantes_mantenimiento_correctivo <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_mantenimiento_correctivo,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=1).tipo
            })

        km_restantes_filtro_aceite = vehiculo.km_restantes_intervalo_cambio_filtro_aceite()
        if km_restantes_filtro_aceite <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aceite,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=3).tipo
            })

        km_restantes_filtro_aire_combustible = vehiculo.km_restantes_intervalo_cambio_filtro_aire_combustible()
        if km_restantes_filtro_aire_combustible <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_aire_combustible,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=4).tipo
            })    

        km_restantes_filtro_caja_corona = vehiculo.km_restantes_intervalo_cambio_aceite_caja_corona()
        if km_restantes_filtro_caja_corona <= km_alerta:
            alertas.append({
                'vehiculo': vehiculo,
                'km_restantes': km_restantes_filtro_caja_corona,
                'tipo': TipoMantenimientoVehiculo.objects.get(id=5).tipo
            })    

    alertas_ordenadas = sorted(alertas, key=lambda x: x['km_restantes'])
    total_alertas = len(alertas_ordenadas)
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': total_alertas,
        'alert_form': alert_form,
    }
    return render(request, 'SGE_vehiculo/alertas.html', context)




@login_required
def tabla_mantenimientos(request):
    vehiculos = Vehiculo.objects.all()
    tipos_mantenimiento = TipoMantenimientoVehiculo.objects.all()
    for vehiculo in vehiculos:
        vehiculo.mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
    context = {
        'vehiculos': vehiculos,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'SGE_vehiculo/tablas.html', context) 




@login_required
def crear_vehiculo(request):
    if request.method == 'GET':
        form = VehiculoForm()
        context = {
            'form': form
        }
        return render(request, 'SGE_vehiculo/nuevo.html', context)
    if request.method == 'POST':
        form = VehiculoForm(request.POST, request.FILES)  # Asegúrate de pasar request.FILES al formulario
        if form.is_valid():
            km_recorridos = form.cleaned_data.get('km_recorridos')
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            intervalo_cambio_filtro_aceite = form.cleaned_data.get('intervalo_cambio_filtro_aceite')
            intervalo_cambio_filtro_aire_combustible = form.cleaned_data.get('intervalo_cambio_filtro_aire_combustible')
            intervalo_cambio_aceite_caja_corona = form.cleaned_data.get('intervalo_cambio_aceite_caja_corona')
            
            if km_recorridos < 1:
                form.add_error('km_recorridos', 'Los Km recorridos no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)
            if intervalo_mantenimiento < 1:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)
            if intervalo_cambio_filtro_aceite  < 1:
                form.add_error('intervalo_cambio_filtro_aceite', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)
            if intervalo_cambio_filtro_aire_combustible < 1:
                form.add_error('intervalo_cambio_filtro_aire_combustible', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)
            if intervalo_cambio_aceite_caja_corona < 1:
                form.add_error('intervalo_cambio_aceite_caja_corona', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'form': form
                }
                return render(request, 'SGE_vehiculo/nuevo.html', context)            
            else:
                if 'imagen' in request.FILES:
                    form.instance.imagen = request.FILES['imagen']
                form.save()
                return redirect('vehiculo')
        else:
            context = {
                'form': form
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo") 
            return render(request, 'SGE_vehiculo/nuevo.html', context)    




@login_required    
def detalles(request, id):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
        form = VehiculoForm(instance=vehiculo)
        context = {
            'vehiculo': vehiculo,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/detalles.html', context)
    
    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        form = VehiculoForm(instance=vehiculo)
        form = VehiculoForm(request.POST, request.FILES, instance=vehiculo)

        if form.is_valid():
            km_recorridos = form.cleaned_data.get('km_recorridos')
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            intervalo_cambio_filtro_aceite = form.cleaned_data.get('intervalo_cambio_filtro_aceite')
            intervalo_cambio_filtro_aire_combustible = form.cleaned_data.get('intervalo_cambio_filtro_aire_combustible')
            intervalo_cambio_aceite_caja_corona = form.cleaned_data.get('intervalo_cambio_aceite_caja_corona')
            if km_recorridos < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('km_recorridos', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            if intervalo_mantenimiento < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            if intervalo_cambio_filtro_aceite < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_cambio_filtro_aceite', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            if intervalo_cambio_filtro_aire_combustible < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_cambio_filtro_aire_combustible', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)
            if intervalo_cambio_aceite_caja_corona < 0:
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                form.add_error('intervalo_cambio_aceite_caja_corona', 'El intervalo de mantenimiento no puede ser un número negativo')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                previous_url = request.META.get('HTTP_REFERER')
                return HttpResponseRedirect(previous_url)            
            else:
                form.save()
                mantenimientos = vehiculo.mantenimientovehiculo_set.all().order_by('-fecha_fin', '-hora_fin')
                context = {
                    'vehiculo': vehiculo,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                }
                return render(request, 'SGE_vehiculo/detalles.html', context) 
        
        else:
            previous_url = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(previous_url) 




@login_required
def eliminar(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    vehiculo.delete()
    return redirect ('vehiculo')    
# fin de vistas generales----------------------------------------------------------------------------------





@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)



@login_required
def mantenimientos_vehiculo(request, id, mant):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)
        mantenimientos = vehiculo.mantenimientovehiculo_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
        context = {
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
            'mantenimientos': mantenimientos,
        }
        return render(request, 'SGE_vehiculo/manteniminetos.html', context)


@login_required
def nuevo_mantenimineto_vehiculo(request, id, mant):
    if request.method == 'GET':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)
        form_mant = MantenimientoVehiculoForm()
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/nuevo_mantenimineto.html', context)

    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)
        form_mant = MantenimientoVehiculoForm(request.POST, request.FILES)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            if 'imagen' in request.FILES:
                mantenimiento.imagen = request.FILES['imagen']
            if form_mant.cleaned_data['km_recorridos'] > vehiculo.km_recorridos:
                form_mant.add_error('km_recorridos', 'Los Km recorridos del mantenimineto no pueden ser mayores que los Km recorridos del vehículo en general')
                context = {
                    'form_mant': form_mant,
                    'vehiculo': vehiculo,
                    'tipo_mantenimiento': tipo_mantenimiento,
                }
                messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")     
                return render(request, 'SGE_vehiculo/nuevo_mantenimineto.html', context)
            else:
                mantenimiento.save()
                return redirect('mantenimientos_vehiculo', id=vehiculo.id, mant=mant)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/nuevo_mantenimineto.html', context)

    return HttpResponse("Method Not Allowed", status=405)        



@login_required
def mod_mantenimineto_vehiculo(request, id, mant):
    if request.method == 'GET':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoForm(instance=mantenimiento)
        context = {
            'form_mant': form_mant,
            'vehiculo': vehiculo,
            'tipo_mantenimiento': tipo_mantenimiento,
        }
        return render(request, 'SGE_vehiculo/mod_mantenimineto.html', context)

    if request.method == 'POST':
        mantenimiento = get_object_or_404(MantenimientoVehiculo, id=id)
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)
        vehiculo = mantenimiento.vehiculo
        form_mant = MantenimientoVehiculoForm(request.POST, request.FILES, instance=mantenimiento)

        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.vehiculo = vehiculo
            mantenimiento.tipo = tipo_mantenimiento
            if 'image' in request.FILES:
                mantenimiento.image = request.FILES['image']
            if form_mant.cleaned_data['km_recorridos'] > vehiculo.km_recorridos:
                form_mant.add_error('km_recorridos', 'Los Km recorridos del mantenimineto no pueden ser mayores que los Km recorridos del vehículo en general')
                context = {
                    'form_mant': form_mant,
                    'vehiculo': vehiculo,
                    'tipo_mantenimiento': tipo_mantenimiento,
                }
                return render(request, 'SGE_vehiculo/mod_mantenimineto.html', context) 
                return redirect('nuevo_mantenimineto_vehiculo', id=vehiculo.id, mant=mant)
            else:
                mantenimiento.save()
                return redirect('mantenimientos_vehiculo', id=vehiculo.id, mant=mant)
        else:
            context = {
                'form_mant': form_mant,
                'vehiculo': vehiculo,
                'tipo_mantenimiento': tipo_mantenimiento,
            }
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'SGE_vehiculo/mod_mantenimineto.html', context)

    return HttpResponse("Method Not Allowed", status=405)




# descargas -----------------------------------------------------------------------------------




@login_required
def documento_general_mantenimientos_vehiculo(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoVehiculo.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_vehiculos_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_vehiculos_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Marca', 'Modelo', 'Tipo', 'Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.vehiculo.marca,
            mantenimiento.vehiculo.modelo,
            mantenimiento.tipo.tipo,
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response





@login_required
def documento_mantenimientos_vehiculo(request, id, mant):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = mant
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, id=mant)

    vehiculo = get_object_or_404(Vehiculo, pk=id)
    mantenimientos = MantenimientoVehiculo.objects.filter(vehiculo=vehiculo).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoVehiculo, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, vehiculo.marca, vehiculo.modelo, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, vehiculo.marca, vehiculo.modelo, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Km Recorridos', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.km_recorridos,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response